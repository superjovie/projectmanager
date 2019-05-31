
from flask import Flask, render_template, request, Response, redirect, url_for, flash, send_from_directory
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os, project, shutil, datetime, json
from setproject import SetProfile
from flask_login import LoginManager
from filemanager import FileManager


fm = FileManager('static/uploads', False)


app = Flask(__name__)



login_manager = LoginManager()
login_manager.init_app(app)

app.secret_key = 'somepasswordisverystrong'
basedir = os.path.abspath(os.path.dirname(__file__))
basepath = 'static/uploads'

@app.route('/list', methods=["POST", "GET"])
def list_():
    s = request.data
    return json.dumps(fm.list(json.loads(s)))

@app.route('/rename', methods=["POST", "GET"])
def rename():
    s = request.data
    return json.dumps(fm.rename(json.loads(s)))

@app.route('/copy', methods=["POST", "GET"])
def copy():
    s = request.data
    return json.dumps(fm.copy(json.loads(s)))

@app.route('/remove', methods=["POST", "GET"])
def remove():
    s = request.data
    return json.dumps(fm.remove(json.loads(s)))

@app.route('/edit', methods=["POST", "GET"])
def edit():
    s = request.data
    return json.dumps(fm.edit(json.loads(s)))

@app.route('/createFolder', methods=["POST", "GET"])
def createFolder():
    s = json.loads(request.get_data())
    return json.dumps(fm.createFolder(s))

@app.route('/changePermissions', methods=["POST", "GET"])
def changePermissions():
    s = request.data
    return json.dumps(fm.changePermissions(json.loads(s)))

@app.route('/compress', methods=["POST", "GET"])
def compress():
    s = request.data
    return json.dumps(fm.compress(json.loads(s)))

@app.route('/downloadMultiple', methods=["POST", "GET"])
def downloadMultiple():
    s = request.values.get('toFilename')
    return fm.downloadMultiple(request.values, Response)
    shutil.rmtree(tmpdir, ignore_errors=True)

@app.route('/move', methods=["POST", "GET"])
def move():
    s = request.data
    return json.dumps(fm.move(json.loads(s)))

@app.route('/getContent', methods=["POST", "GET"])
def getContent():
    s = request.data
    return json.dumps(fm.getContent(json.loads(s)))

@app.route('/extract', methods=["POST", "GET"])
def extract():
    s = request.data
    return json.dumps(fm.extract(json.loads(s)))

@app.route('/upload', methods=["POST", "GET"])
def upload():
    return json.dumps(fm.upload(request.files, request.form['destination']))

@app.route("/download", methods=['POST', 'GET'])  # 下载项目文件
def download() :
    s = request.values['path']
    return fm.download(request.values['path'], Response)






@app.route('/getcount/<co>/<dir>', methods=["POST", "GET"])
def getcount(co,dir) :
    a = []
    company = basedir + '/static/uploads/' + co
    d = 0
    if dir is '' :
        for dirs in os.listdir(company) :
            a.append(dirs)

        try :
            a.remove('super.xlsx')
        except :
            print('directory error')
        try :
            a.remove('.DS_Store')
        except :
            print('directory error')

        for dirss in a :
            c = os.listdir(company + '/' + dirss)
            try :
                c.remove('.DS_Store')
                d += len(c)
            except :
                d += len(c)

    if dir is not '' :
        c = os.listdir(company+'/'+dir)
        try :
            c.remove('.DS_Store')
            d += len(c)
        except :
            d += len(c)
    msg = {
        "value" : str(d)
    }
    return json.dumps(msg, ensure_ascii=False)

@app.route('/allmoney/<co>/<dir>', methods=["POST", "GET"])
def allmoney(co,dir) :
    value = 0
    company = basedir + '/static/uploads/' + co
    if dir is '' :
        for dirs in os.listdir(company) :
            if dirs == '.DS_Store' :
                continue
            if dirs == 'super.xlsx':
                continue
            for dirss in os.listdir(company + '/' + dirs) :
                if dirss == '.DS_Store' :
                    continue
                wb = load_workbook(company + '/' + dirs + '/' + dirss + '/' + '项目情况表.xlsx')
                sheet = wb['Sheet1']
                if sheet['H2'].value is not None :
                    num = int(sheet['H2'].value)
                    value = value + num

    if dir is not '' :
        for dirs in os.listdir(company + '/' + dir) :
            if dirs == '.DS_Store' :
                continue
            if dirs == 'super.xlsx':
                continue
            wb = load_workbook(company + '/' + dir + '/' + dirs + '/' + '项目情况表.xlsx')
            sheet = wb['Sheet1']
            if sheet['H2'].value is not None :
                num = int(sheet['H2'].value)
                value = value + num
    msg = {
        "value" : value
    }
    return json.dumps(msg, ensure_ascii=False)

@app.route('/getendcount/<co>/<dir>', methods=["POST", "GET"])
def getendcount(co,dir) :
    c = 0
    company = basedir + '/static/uploads/' + co
    wb = load_workbook(company + '/super.xlsx')
    if dir is '' :
        for sheets in wb :
            for i in range(3, 200) :
                sheet = wb[sheets.title]
                if sheet['B' + str(i)].value is not None and sheet['P' + str(i)].value == '完工' :
                    c = c + 1
                if sheet['B' + str(i)].value is None and sheet['P' + str(i)].value is None :
                    break
    if dir is not '' :
        sheet = wb[dir]
        for i in range(3, 200) :
            if sheet['B' + str(i)].value is not None and sheet['P' + str(i)].value == '完工' :
                c = c + 1
            if sheet['B' + str(i)].value is None and sheet['P' + str(i)].value is None :
                break
    msg = {
        "value" : str(c)
    }
    return json.dumps(msg, ensure_ascii=False)

@app.route('/getallmoney/<dir>',methods=['POST','GET'])
def getallmoney(dir):
    #wb = load_workbook(basepath + '/' + dir+'/super.xlsx')
    c = 0
    m = 0
    e = 0

    count = json.loads(getcount(dir,''))
    if count is not None:c = c + int(count['value'])
    else:c = c + 0
    money = json.loads(allmoney(dir,''))
    if money is not None:m = m + int(money['value'])
    else:m = m + 0
    endcount = json.loads(getendcount(dir,''))
    if endcount is not None:e = e + int(endcount['value'])
    else:e = e + 0
    msg = {
        "count" : c,
        "money" : m,
        "endcount" : e
    }
    return json.dumps(msg, ensure_ascii=False)




@app.route('/getyear/<dir>', methods=["POST", "GET"])
def getyear(dir) :
    a = []
    company = 'static/uploads/'+dir
    for dirs in os.listdir(company) :
        a.append(dirs)
        try :
            a.remove('.DS_Store')
        except :
            continue
    msg = {
        'year' : a
    }
    return json.dumps(msg, ensure_ascii=False)


@app.route('/loadcompany', methods=["POST", "GET"])
def loadcompany() :
    a = []
    company = 'static/uploads'
    for dirs in os.listdir(company) :
        a.append(dirs)
        try :
            a.remove('.DS_Store')
        except :
            continue
    msg = {
        'company' : a
    }
    return json.dumps(msg, ensure_ascii=False)


@app.route('/getname/<co>/<year>', methods=["POST", "GET"])
def getname(co,year) :
    company = 'static/uploads/' + co + '/'
    a = []
    for dirs in os.listdir(company + year) :
        a.append(dirs)
        try :
            a.remove('.DS_Store')
        except :
            continue
    msg = {
        'name' : a
    }
    return json.dumps(msg, ensure_ascii=False)


@app.route('/getalert/<dir>', methods=['POST', 'GET'])
def getalert(dir) :
    wb = load_workbook(basepath+'/'+dir+'/super.xlsx', data_only=True)
    type = []
    msg = []
    name = []
    for sheets in wb :
        sheet = wb[sheets.title]
        for i in range(3, 200) :
            if sheet['B' + str(i)].value is not None and sheet['p' + str(i)].value != '完工' :
                value = sheet['F' + str(i)].value

                values = sheet['M' + str(i)].value

                if value is not None :
                    value = str(value)
                    strdata1 = value.split(' ')
                    date1 = strdata1[0].replace('-', '/')
                    planend = datetime.datetime.strptime(date1, '%Y/%m/%d')
                    now = datetime.datetime.now()
                    between = planend - now
                    if 0 < between.days < 7 :
                        name.append(sheet['B' + str(i)].value)
                        type.append("工期")
                        msg.append(str(between.days))

                    if between.days <= 0 :
                        name.append(sheet['B' + str(i)].value)
                        type.append("工期")
                        msg.append(str(between.days))

                if values is not None :
                    values = str(values)
                    strdata2 = values.split(' ')
                    date2 = strdata2[0].replace('-', '/')
                    planins = datetime.datetime.strptime(date2, '%Y/%m/%d')
                    now = datetime.datetime.now()
                    betweens = planins - now
                    if 0 < betweens.days < 7 :
                        name.append(sheet['B' + str(i)].value)
                        msg.append(str(betweens.days))
                        type.append("保险")
                    if betweens.days <= 0 :
                        name.append(sheet['B' + str(i)].value)
                        msg.append(str(betweens.days))
                        type.append("保险")

            if sheet['B' + str(i)].value is None :
                break
    msgs = {
        "name" : name,
        "msg" : msg,
        "type" : type
    }

    return json.dumps(msgs, ensure_ascii=False)


@app.route('/getallname/<co>',methods=['POST','GET'])
def getallname(co):
    allname = []
    allyear = []
    company = 'static/uploads/'+co
    wb = load_workbook(company + '/super.xlsx')
    for sheets in wb:
        sheet = wb[sheets.title]
        i = 3
        while 1:
            if sheet['B'+str(i)].value is not None:
                allname.append(sheet['B'+str(i)].value)
                allyear.append(sheets.title)
                i = i + 1
            else:
                break
        msg = {
            "allname":allname,
            "allyear":allyear
        }
    return json.dumps(msg,ensure_ascii=False)


@app.route('/initall')
def initall() :
    wb = load_workbook('super.xlsx')
    for sheets in wb :
        if not os.path.exists('static/uploads/' + sheets.title) :
            os.mkdir('static/uploads/' + sheets.title)
        sheet = wb[sheets.title]
        dirlist = []
        for i in range(3, 200) :
            if sheet['B' + str(i)].value is not None :
                dirlist.append(sheet['B' + str(i)].value)
            if sheet['B' + str(i)].value is None :
                break
        for dir in dirlist :
            if not os.path.exists('static/uploads/' + sheets.title + '/' + dir) :
                os.mkdir('static/uploads/' + sheets.title + '/' + dir)
            if not os.path.exists('static/uploads/' + sheets.title + '/' + dir + '/项目情况表.xlsx') :
                shutil.copyfile('项目情况表.xlsx', 'static/uploads/' + sheets.title + '/' + dir + '/项目情况表.xlsx')
    return render_template("index.html")


@app.route('/uploader/<co>', methods=['POST', 'GET'])
def uploader(co) :
    year = request.values.get('year')
    name = request.values.get('name')
    path = request.values.get('path')
    global upload_path
    file = request.files['file']
    folder = os.path.join('static/uploads/' + co + '/' + year + '/' + name + '/',
                          path)
    if not os.path.exists(folder) :
        os.mkdir(folder)
    basepath = os.path.dirname(__file__)  # 当前文件所在路径

    upload_path = os.path.join(basepath, folder, secure_filename(file.filename))  # 注意：没有的文件夹一定要先创建，不然会提示没有该路径

    file.save(upload_path)
    msg = {
        "code" : 0
        , "msg" : ""
        , "data" : {
            "src" : "http://cdn.layui.com/123.jpg"
        }
    }
    return json.dumps(msg, ensure_ascii=False)


@app.route('/showproject/<com>/<year>/<name>',methods=['POST','GET'])
def showproject(com,year,name):
    msg = json.loads(getproject(com,year,name))
    datalist = msg["data"]
    len1 = len(datalist[13])
    len2 = len(datalist[15])
    len3 = len(datalist[19])
    len4 = len(datalist[21])
    ra =  render_template('main.html',x=datalist,year=year,len1=len1,len2=len2,len3=len3,len4=len4)
    msg={"data":ra}
    return json.dumps(msg, ensure_ascii=False)


@app.route('/getproject/<com>/<year>/<name>', methods=['POST', 'GET'])
def getproject(com,year,name) :
    data = []
    if request.method == 'POST' :
        pro = project.Project(com,'static/uploads/' + com + '/' + year + '/' + name + '/项目情况表.xlsx', year)
        data.append(pro.Name)  # 0
        data.append(pro.Address)  # 1
        data.append(pro.ContractDays)  # 2
        data.append(pro.ContractContent)  # 3
        data.append(pro.BiddingTime)  # 4
        data.append(pro.OpenTime)  # 5
        data.append(pro.LockPerson)  # 6
        data.append(pro.PayMethod)  # 7
        data.append(pro.Owner)  # 8
        data.append(pro.OwnerContact)  # 9
        data.append(pro.Worker)  # 10
        data.append(pro.WorkerContact)  # 11
        data.append(pro.Total)  # 12
        data.append(pro.ProfileName)  # 13
        data.append(pro.ProfileContent)  # 14
        data.append(pro.PatrolTime)  # 15
        data.append(pro.PatrolMan)  # 16
        data.append(pro.PatrolReason)  # 17
        data.append(pro.PatrolResuilt)  # 18
        data.append(pro.IndexPaymentTime)  # 19
        data.append(pro.IndexPayment)  # 20
        data.append(pro.RemarkTime)  # 21
        data.append(pro.RemarkReason)  # 22
        data.append(pro.RemarkResuilts)  # 23
        data.append(pro.Insurance)  # 24
        data.append(pro.InsuranceEnd)  # 25
        data.append(pro.InsuranceExpire)  # 26
        data.append(pro.InsuranceFile)  # 27
        data.append(pro.Endding)  # 28
        data.append(pro.PlanEnd)  # 29
        data.append(pro.NotEnddingRemark)  # 30
        del pro
    msg = {
        'data' : data
    }

    del data
    return json.dumps(msg, ensure_ascii=False)


@app.route('/set/<com>', methods=['POST', 'GET'])
def set(com) :
    datalist = []
    year = ''
    if request.method == 'POST' :
        year = request.values.get('quiz2')

        if request.values.get('Name') is not None :
            datalist.append(request.values.get('Name'))  # 0
        else :
            datalist.append('')
        if request.values.get('Address') is not None :
            datalist.append(request.values.get('Address'))  # 1
        else :
            datalist.append('')
        if request.values.get('ContractDays') is not None :
            datalist.append(request.values.get('ContractDays'))  # 2
        else :
            datalist.append('')
        if request.values.get('ContractContent') is not None :
            datalist.append(request.values.get('ContractContent'))  # 3
        else :
            datalist.append('')
        if request.values.get('BiddingTime') is not None :
            datalist.append(request.values.get('BiddingTime'))  # 4
        else :
            datalist.append('')
        if request.values.get('OpenTime') is not None :
            datalist.append(request.values.get('OpenTime'))  # 5
        else :
            datalist.append('')
        if request.values.get('LockPerson') is not None :
            datalist.append(request.values.get('LockPerson'))  # 6
        else :
            datalist.append('')
        if request.values.get('PayMethod') is not None :
            datalist.append(request.values.get('PayMethod'))  # 7
        else :
            datalist.append('')
        if request.values.get('Owner') is not None :
            datalist.append(request.values.get('Owner'))  # 8
        else :
            datalist.append('')
        if request.values.get('OwnerContact') is not None :
            datalist.append(request.values.get('OwnerContact'))  # 9
        else :
            datalist.append('')
        if request.values.get('Worker') is not None :
            datalist.append(request.values.get('Worker'))  # 10
        else :
            datalist.append('')
        if request.values.get('WorkerContact') is not None :
            datalist.append(request.values.get('WorkerContact'))  # 11
        else :
            datalist.append('')
        if request.values.getlist('P[]') is not None :
            datalist.append(request.values.getlist('P[]'))  # 12
        else :
            datalist.append('')
        da = request.values
        if request.values.getlist('F[]') is not None :
            datalist.append(request.values.getlist('F[]'))  # 13
        else :
            datalist.append('')
        if request.values.getlist('Pt[]') is not None :
            datalist.append(request.values.getlist('Pt[]'))  # 14
        else :
            datalist.append('')
        if request.values.getlist('Pn[]') is not None :
            datalist.append(request.values.getlist('Pn[]'))  # 15
        else :
            datalist.append('')
        if request.values.getlist('Pr[]') is not None :
            datalist.append(request.values.getlist('Pr[]'))  # 16
        else :
            datalist.append('')
        if request.values.getlist('Prs[]') is not None :
            datalist.append(request.values.getlist('Prs[]'))  # 17
        else :
            datalist.append('')
        if request.values.getlist('PayTime[]') is not None :
            datalist.append(request.values.getlist('PayTime[]'))  # 18
        else :
            datalist.append('')
        if request.values.getlist('Payment[]') is not None :
            datalist.append(request.values.getlist('Payment[]'))  # 19
        else :
            datalist.append('')
        if request.values.getlist('RemarkTime[]') is not None :
            datalist.append(request.values.getlist('RemarkTime[]'))  # 20
        else :
            datalist.append('')
        if request.values.getlist('RemarkReason[]') is not None :
            datalist.append(request.values.getlist('RemarkReason[]'))  # 21
        else :
            datalist.append('')
        if request.values.getlist('RemarkResuilt[]') is not None :
            datalist.append(request.values.getlist('RemarkResuilt[]'))  # 22
        else :
            datalist.append('')
        if request.values.get('Insurance') is not None :
            datalist.append(request.values.get('Insurance'))  # 23
        else :
            datalist.append('')
        if request.values.get('InsuranceEnd') is not None :
            datalist.append(request.values.get('InsuranceEnd'))  # 24
        else :
            datalist.append('')
        datalist.append('')                                     # 25
        if request.values.get('InsuranceFile') is not None :
            datalist.append(request.values.get('InsuranceFile'))  # 26
        else :
            datalist.append('')
        if request.values.get('Endding') is not None :
            datalist.append(request.values.get('Endding'))  # 27
        else :
            datalist.append('')
        if request.values.get('NotEnddingRemark') is not None :
            datalist.append(request.values.get('NotEnddingRemark'))  # 28
        else :
            datalist.append('')
        if request.values.get('Total') is not None :
            datalist.append(request.values.get('Total'))  # 29
        else :
            datalist.append('')
        SetProfile(com,year, datalist)
    msg = {
        'data' : datalist,
        'year' : year
    }
    return json.dumps(msg, ensure_ascii=False)


@app.route('/')
def main() :
    return render_template('index.html')

@app.route('/manager')
def manager():
    return render_template('NewPro.html')



if __name__ == '__main__' :
    app.run(host='0.0.0.0', port=80)
