from openpyxl import load_workbook, workbook
import datetime, os, shutil


class SetProfile :

    def __init__(self,com,year, datalist) :
        planend = ''
        dir = 'static/uploads/' + com + '/' +year + '/' + datalist[0]
        if not os.path.exists(dir) :
            os.mkdir(dir)
            shutil.copyfile('项目情况表.xlsx', dir + '/项目情况表.xlsx')

        wb = load_workbook(dir + "/" + '项目情况表.xlsx', data_only=True)
        sheet = wb['Sheet1']
        if datalist[29] is not ''  :
            sheet['H2'].value = datalist[29]
        if datalist[0] is not ''  :
            sheet['E2'].value = datalist[0]
        if datalist[1] is not '' :
            sheet['E3'].value = datalist[1]
        if datalist[2] is not '' :
            sheet['H3'].value = datalist[2]
        if datalist[3] is not '' :
            sheet['E4'].value = datalist[3]
        if datalist[4] is not '' :
            sheet['E5'].value = datalist[4]
        if datalist[5] is not '' :
            sheet['G5'].value = datalist[5]
        if datalist[6] is not '' :
            sheet['E6'].value = datalist[6]
        if datalist[7] is not '' :
            sheet['G6'].value = datalist[7]
        if datalist[8] is not '' :
            sheet['E7'].value = datalist[8]
        if datalist[9] is not '' :
            sheet['G7'].value = datalist[9]
        if datalist[10] is not '' :
            sheet['E8'].value = datalist[10]
        if datalist[11] is not '' :
            sheet['G8'].value = datalist[11]
        sheet['B9'].value = '计划竣工时间'

        if datalist[12] is not '' :
            for p in range(0,len(datalist[12])):
                for i in range(9, 18) :  # 资料信息
                    if sheet['B' + str(i)].value == datalist[12][p] :
                        sheet['E' + str(i)].value = datalist[13][p]
                        break
                    elif sheet['F' + str(i)].value == datalist[12][p] :
                        sheet['G' + str(i)].value = datalist[13][p]
                        break

                    elif sheet['B' + str(i)].value is None :
                        sheet['B' + str(i)].value = datalist[12][p]
                        sheet['E' + str(i)].value = datalist[13][p]
                        break
                    elif sheet['F' + str(i)].value is None :
                        sheet['F' + str(i)].value = datalist[12][p]
                        sheet['G' + str(i)].value = datalist[13][p]
                        break

        if sheet['G5'].value is not '' and None :
            strdata = str(sheet['G5'].value)
            strdata = strdata.split(' ')
            strs = strdata[0].replace('-', '/')
            planenddate = datetime.datetime.strptime(strs, '%Y/%m/%d') + datetime.timedelta(days=int(sheet['H3'].value))
            planend = planenddate.strftime('%Y/%m/%d')
            sheet['E9'].value = planend

        if datalist[14] is not '' :
            i = 19
            for p in range(0,len(datalist[14])):# 巡查
                sheet['C' + str(i)].value = datalist[14][p]
                sheet['D' + str(i)].value = datalist[15][p]
                sheet['E' + str(i)].value = datalist[16][p]
                sheet['F' + str(i)].value = datalist[17][p]
                i = i + 1

        if datalist[18] is not '' :
            p=19# 付款
            for i in range(0, len(datalist[18])) :
                sheet['H' + str(p)].value = datalist[18][i]
                sheet['I' + str(p)].value = datalist[19][i]
                p = p + 1

        if datalist[20] is not '' :
            p=28# 备注时间
            for i in range(0, len(datalist[20])) :
                    sheet['C' + str(p)].value = datalist[20][i]
                    sheet['D' + str(p)].value = datalist[21][i]
                    sheet['F' + str(p)].value = datalist[22][i]
                    p = p + 1


        wb.save('static/uploads/' + com + '/' + year + "/" + datalist[0] + "/" + '项目情况表.xlsx')
        superwb = load_workbook('static/uploads/'+com+'/super.xlsx', data_only=True)
        supersheet = superwb[year]
        for i in range(3, 100) :
            if supersheet['B' + str(i)].value is not None:
                projectile = supersheet['B' + str(i)].value
                if projectile.find(datalist[0]) >= 0 :

                    if datalist[23] is not '' :
                        supersheet['L' + str(i)].value = datalist[23]
                    if datalist[24] is not '' :
                        supersheet['M' + str(i)].value = datalist[24]
                        strdata = str(supersheet['M' + str(i)].value)
                        strdata = strdata.split(' ')
                        strs = strdata[0].replace('-', '/')
                        expire = datetime.datetime.strptime(strs, '%Y/%m/%d') - datetime.datetime.now()
                        supersheet['N' + str(i)].value = int(expire.days)
                    if datalist[26] is not '' :
                        supersheet['O' + str(i)].value = datalist[26]
                    if datalist[27] is '1' :
                        supersheet['P' + str(i)].value = '完工'
                    else:supersheet['P' + str(i)].value = '/'
                    supersheet['F' + str(i)].value = planend
                    if datalist[28] is not '' :
                        supersheet['Q' + str(i)].value = datalist[28]
                    wbs = load_workbook('static/uploads/' + com + '/' + year + "/" + datalist[0] + "/" + '项目情况表.xlsx')
                    sheets = wbs['Sheet1']
                    supersheet['C' + str(i)].value = sheets['E5'].value
                    supersheet['D' + str(i)].value = sheets['E8'].value
                    supersheet['E' + str(i)].value = sheets['G5'].value
                    supersheet['J' + str(i)].value = sheets['E13'].value  ##F竣工时间
                    if sheets['H3'].value is not '' and None :
                        supersheet['G' + str(i)].value = int(sheets['H3'].value)
                    break
            else :
                supersheet['B' + str(i)].value = datalist[0]
                if datalist[23] is not '' :
                    supersheet['L' + str(i)].value = datalist[23]
                if datalist[24] is not '' :
                    supersheet['M' + str(i)].value = datalist[24]
                    strdata = str(supersheet['M' + str(i)].value)
                    strdata = strdata.split(' ')
                    strs = strdata[0].replace('-', '/')
                    expire = datetime.datetime.strptime(strs, '%Y/%m/%d') - datetime.datetime.now()
                    supersheet['N' + str(i)].value = int(expire.days)
                if datalist[26] is not '' :
                    supersheet['O' + str(i)].value = datalist[26]
                if datalist[27] is '1' :
                    supersheet['P' + str(i)].value = '完工'
                else:supersheet['P' + str(i)].value = '/'
                supersheet['F' + str(i)].value = planend
                if datalist[28] is not '' :
                    supersheet['Q' + str(i)].value = datalist[28]
                wbs = load_workbook('static/uploads/' + year + "/" + datalist[0] + "/" + '项目情况表.xlsx')
                sheets = wbs['Sheet1']
                supersheet['C' + str(i)].value = sheets['E5'].value
                supersheet['D' + str(i)].value = sheets['E8'].value
                supersheet['E' + str(i)].value = sheets['G5'].value
                supersheet['J' + str(i)].value = sheets['E13'].value  #F竣工时间
                if sheets['H3'].value is not '' and None :
                    supersheet['G' + str(i)].value = int(sheets['H3'].value)
                break
        superwb.save('static/uploads/'+com+'/super.xlsx')
