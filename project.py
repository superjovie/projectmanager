from openpyxl import load_workbook


class Project:
    Name = ''  # 项目名称
    Adress = ''  # 项目地址
    ContractDays = ''  # 合同工期
    ContractContent = ''  # 项目内容
    BiddingTime = ''  # 中标时间
    OpenTime = ''  # 开工时间
    LockPerson = ''  # 锁人信息
    PayMethod = ''  # 付款方式
    Owner = ''  # 业主/业人代表
    OwnerContact = ''  # 业人联系方式
    Worker = ''  # 项目联系人
    WorkerContact = ''  # 项目联系人联系方式
    Total = 0  # 合同总金额

    Remain = 0  # 合同剩余未付款金额
    sub = 0  # 合同已付金额
    IndexPayment = []  # 付款
    IndexPaymentTime = []  # 付款日期

    RemarkReason = []  # 备注原因
    RemarkResuilts = []  # 备注结果
    RemarkTime = []  # 备注时间

    PatrolTime = []  # 巡查时间
    PatrolMan = []  # 巡查人
    PatrolReason = []  # 巡查原因
    PatrolResuilt = []  # 巡查结果

    ProfileName = []  # 项目资料信息名字
    ProfileContent = []  # 项目资料信息内容

    EndTime = ''  # 竣工时间
    Insurance = ''  # 保险
    InsuranceFile = ''  # 保险原件/扫描件
    InsuranceEnd = ''  # 保险到期日
    InsuranceExpire = ''  # 保险超期天数

    Endding = False  # 是否完工
    NotEnddingRemark = ''  # 未完工备注
    PlanEnd = ''  #计划完工时间
    def datetostr(self,data):
        data = str(data)
        data = data.split(' ')
        strs = data[0].replace('-', '/')
        return strs

    def __init__(self,com, name, year):
        self.Name = ''  # 项目名称
        self.Adress = ''  # 项目地址
        self.ContractDays = ''  # 合同工期
        self.ContractContent = ''  # 项目内容
        self.BiddingTime = ''  # 中标时间
        self.OpenTime = ''  # 开工时间
        self.LockPerson = ''  # 锁人信息
        self.PayMethod = ''  # 付款方式
        self.Owner = ''  # 业主/业人代表
        self.OwnerContact = ''  # 业人联系方式
        self.Worker = ''  # 项目联系人
        self.WorkerContact = ''  # 项目联系人联系方式
        self.Total = 0  # 合同总金额

        self.Remain = 0  # 合同剩余未付款金额
        self.sub = 0  # 合同已付金额
        self.IndexPayment = []  # 付款
        self.IndexPaymentTime = []  # 付款日期

        self.RemarkReason = []  # 备注原因
        self.RemarkResuilts = []  # 备注结果
        self.RemarkTime = []  # 备注时间

        self.PatrolTime = []  # 巡查时间
        self.PatrolMan = []  # 巡查人
        self.PatrolReason = []  # 巡查原因
        self.PatrolResuilt = []  # 巡查结果

        self.ProfileName = []  # 项目资料信息名字
        self.ProfileContent = []  # 项目资料信息内容

        self.EndTime = ''  # 竣工时间
        self.Insurance = ''  # 保险
        self.InsuranceFile = ''  # 保险原件/扫描件
        self.InsuranceEnd = ''  # 保险到期日
        self.InsuranceExpire = ''  # 保险超期天数

        self.Endding = False  # 是否完工
        self.NotEnddingRemark = ''  # 未完工备注
        self.PlanEnd = ''  # 计划完工时间

        wb = load_workbook(name,data_only=True)
        sheet = wb['Sheet1']
        self.PlanEnd = self.datetostr(sheet['E9'].value)

        self.Name = sheet['E2'].value
        self.Address = sheet['E3'].value
        self.ContractDays = sheet['H3'].value
        self.ContractContent = sheet['E4'].value
        self.BiddingTime = self.datetostr(sheet['E5'].value)
        self.OpenTime = self.datetostr(sheet['G5'].value)
        self.LockPerson = sheet['E6'].value
        self.PayMethod = sheet['G6'].value
        self.Owner = sheet['E7'].value
        self.OwnerContact = sheet['G7'].value
        self.Worker = sheet['E8'].value
        self.WorkerContact = sheet['G8'].value
        if sheet['H2'].value is not None:
            self.Total = int(sheet['H2'].value)
        for i in range(9, 18):  # 资料信息
            if sheet['B' + str(i)].value is not None:
                self.ProfileName.append(sheet['B' + str(i)].value)

                self.ProfileContent.append(sheet['E' + str(i)].value)
            if sheet['F' + str(i)].value is not None:
                self.ProfileContent.append(sheet['G' + str(i)].value)
                self.ProfileName.append(sheet['F' + str(i)].value)

        for i in range(19, 27):
            if sheet['C' + str(i)].value is not None:  # 巡查
                self.PatrolMan.append(sheet['D' + str(i)].value)
                self.PatrolReason.append(sheet['E' + str(i)].value)
                self.PatrolTime.append(self.datetostr(sheet['C' + str(i)].value))
                self.PatrolResuilt.append(sheet['F' + str(i)].value)
            if sheet['H' + str(i)].value is not None:  # 付款
                self.IndexPayment.append(int(sheet['I' + str(i)].value))
                self.IndexPaymentTime.append(self.datetostr(sheet['H' + str(i)].value))
                for money in self.IndexPayment:
                    self.sub = self.sub + int(money)
                    self.Remain = self.Total - self.sub

        for i in range(28, 34):
            if sheet['C' + str(i)].value is not None:
                self.RemarkTime.append(self.datetostr(sheet['C' + str(i)].value))
                self.RemarkReason.append(sheet['D' + str(i)].value)
                self.RemarkResuilts.append(sheet['F' + str(i)].value)


        superwb = load_workbook('static/uploads/'+com+'/super.xlsx',data_only=True)
        supersheet = superwb[year]

        for i in range(3,100):
            if supersheet['B' + str(i)].value is not None:

                projectile = supersheet['B' + str(i)].value

                if name.find(projectile) >= 0:
                    self.Insurance = supersheet['L' + str(i)].value
                    self.InsuranceEnd = self.datetostr(supersheet['M' + str(i)].value)
                    self.InsuranceExpire = supersheet['N' + str(i)].value
                    self.InsuranceFile = supersheet['O' + str(i)].value

                    self.EndTime = self.datetostr(supersheet['F' + str(i)].value)
                    self.Endding = supersheet['P' + str(i)].value
                    self.NotEnddingRemark = supersheet['Q' + str(i)].value

                else:
                    continue
            else:
                break

